"""Diagnostico de qualidade da fonte operacional — OBSERVACIONAL.

Le a base operacional mensal (Excel de 5 abas ou pasta com uma CSV por aba) e
descreve o que existe: esquema, contagens, faltantes, duplicidades, chaves sem
correspondencia, valores/status contraditorios e avisos de fonte.

O que este modulo NAO faz (por decisao de metodo — o tratamento e uma mudanca
posterior, com regra aprovada pelo negocio):
  - nao normaliza identificadores, nao deduplica, nao exclui e nao imputa;
  - nao altera o arquivo de origem (abertura sempre somente leitura);
  - nao calcula receita, margem, ranking, classificacao ou qualquer indicador;
  - nao promove anomalia observada a regra aprovada — apenas registra e aponta
    a decisao pendente.

A saida separa o que exige atencao (anomalia/aviso, codigos D-###) do perfil e
inventario da fonte (informativo, codigos P-###), com contadores independentes:
inventario de status nao infla a contagem de problemas.

Textos exibidos ao usuario (titulos, descricoes, decisoes e mensagens) usam
pt-BR acentuado em UTF-8. Chaves do JSON e valores de enumeracao (classe,
severidade, status_evidencia, forca) seguem ASCII de proposito: sao rotulos
tecnicos usados em comparacoes e filtros, nao texto de leitura.

Saidas deterministicas: mesmas entradas produzem bytes identicos (nao ha
carimbo de tempo no conteudo; a auditoria se faz pelo SHA-256 das entradas).

Uso:
    python src/diagnostico_fonte.py --entrada <arquivo.xlsx|pasta-csv> \
        [--saida outputs/diagnostico] [--rotulo 2026-01] [--periodo 2026-01:2026-03]

A leitura de .xlsx exige openpyxl, declarado em requirements.txt com versao
fixa e hash verificado. A leitura de CSVs usa apenas a biblioteca padrao.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from pathlib import Path

VERSAO = "1.0.0"

# ---------------------------------------------------------------------------
# Contrato da fonte: ESTRUTURA esperada (nao regra de negocio).
# Deriva de .project/DATA_CATALOG.md. Vale para qualquer arquivo mensal com a
# mesma estrutura; ausencias e excedentes viram achado, nao erro fatal.
# ---------------------------------------------------------------------------
CONTRATO = {
    "clientes": {
        "aba": "Clientes",
        "arquivo": "clientes.csv",
        "chave": ["cliente_id"],
        "colunas": ["cliente_id", "razao_social", "regiao", "segmento", "canal", "status"],
        "numericas": [],
        "datas": [],
        "categoricas": ["regiao", "segmento", "canal", "status"],
        "essenciais": ["cliente_id"],
    },
    "vendas": {
        "aba": "Vendas",
        "arquivo": "vendas.csv",
        "chave": ["pedido_id"],
        "colunas": [
            "pedido_id", "atualizado_em", "data_pedido", "cliente_id",
            "receita_bruta", "desconto", "custo_produto", "status_pedido",
        ],
        "numericas": ["receita_bruta", "desconto", "custo_produto"],
        "datas": ["atualizado_em", "data_pedido"],
        "categoricas": ["status_pedido"],
        "essenciais": ["pedido_id", "cliente_id", "receita_bruta", "custo_produto"],
    },
    "custos_logisticos": {
        "aba": "Custos_Logisticos",
        "arquivo": "custos_logisticos.csv",
        "chave": ["pedido_id"],
        "colunas": ["pedido_id", "frete", "custo_manuseio"],
        "numericas": ["frete", "custo_manuseio"],
        "datas": [],
        "categoricas": [],
        "essenciais": ["pedido_id", "frete"],
    },
    "visitas": {
        "aba": "Visitas",
        "arquivo": "visitas.csv",
        "chave": ["visita_id"],
        "colunas": ["visita_id", "cliente_id", "mes_ref", "data_planejada", "data_realizada", "status"],
        "numericas": [],
        "datas": ["data_planejada", "data_realizada"],
        "categoricas": ["status"],
        "essenciais": ["visita_id", "cliente_id"],
    },
    "parametros": {
        "aba": "Parametros",
        "arquivo": "parametros.csv",
        "chave": ["parametro"],
        "colunas": ["parametro", "valor", "unidade", "fonte", "status"],
        "numericas": ["valor"],
        "datas": [],
        "categoricas": ["status"],
        "essenciais": ["parametro", "valor"],
    },
}

# Relacionamentos declarados (estrutura). "obrigatoria": a ausencia do destino
# e uma quebra; "informativa": a ausencia e apenas reportada.
RELACIONAMENTOS = [
    {"origem": ["vendas", "cliente_id"], "destino": ["clientes", "cliente_id"], "forca": "obrigatoria"},
    {"origem": ["custos_logisticos", "pedido_id"], "destino": ["vendas", "pedido_id"], "forca": "obrigatoria"},
    {"origem": ["visitas", "cliente_id"], "destino": ["clientes", "cliente_id"], "forca": "obrigatoria"},
    {"origem": ["vendas", "pedido_id"], "destino": ["custos_logisticos", "pedido_id"], "forca": "informativa"},
]

# Decisão pendente compartilhada: D-003 (identificador fora do padrão) e
# D-006 (chave que só casaria após normalização) são achados DIFERENTES,
# preservados separadamente, mas apontam para a MESMA decisão de negócio.
# Usar a mesma formulação canônica evita duplicidade no resumo consolidado.
DECISAO_NORMALIZACAO_IDS = ("O negócio aprova a regra de normalização de identificadores antes de qualquer cruzamento")

RE_INTEIRO = re.compile(r"^-?\d+$")
RE_DECIMAL = re.compile(r"^-?\d+[.,]\d+$")
RE_DATA = re.compile(r"^\d{4}-\d{2}-\d{2}$")
RE_DATAHORA = re.compile(r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}(:\d{2})?$")
RE_COMPETENCIA = re.compile(r"^\d{4}-\d{2}$")
RE_ID_PADRAO = re.compile(r"^[A-Z0-9_.\-]+$")


# ---------------------------------------------------------------------------
# Leitura (somente leitura, sem transformar valores)
# ---------------------------------------------------------------------------
def _texto(valor) -> str:
    """Converte celula em texto SEM limpar conteudo (espacos preservados)."""
    if valor is None:
        return ""
    if hasattr(valor, "isoformat"):
        texto = valor.isoformat(sep=" ") if hasattr(valor, "hour") else valor.isoformat()
        return texto[:16] if getattr(valor, "hour", 0) or getattr(valor, "minute", 0) else texto
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor)


def ler_csvs(pasta: Path) -> tuple[dict, dict, list]:
    tabelas, fontes, achados = {}, {}, []
    for nome, contrato in CONTRATO.items():
        caminho = pasta / contrato["arquivo"]
        if not caminho.exists():
            achados.append(_achado("FONTE", "tabela", nome, "aviso",
                                   f"Tabela esperada ausente na fonte (arquivo {contrato['arquivo']})",
                                   "observado", "O negócio confirma se a tabela foi descontinuada ou se o arquivo veio incompleto"))
            continue
        with caminho.open(newline="", encoding="utf-8") as fh:
            linhas = list(csv.reader(fh))
        cabecalho = linhas[0] if linhas else []
        registros = [dict(zip(cabecalho, linha + [""] * (len(cabecalho) - len(linha)))) for linha in linhas[1:]]
        tabelas[nome] = {"colunas": cabecalho, "registros": registros}
        fontes[nome] = {"origem": contrato["arquivo"], "sha256": _sha256(caminho)}
    return tabelas, fontes, achados


def ler_xlsx(caminho: Path) -> tuple[dict, dict, list]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - ambiente sem openpyxl
        raise SystemExit(
            "ERRO: a leitura de .xlsx exige a biblioteca openpyxl. Instale com "
            "'python -m pip install --require-hashes -r requirements.txt' ou aponte "
            "--entrada para a pasta com as CSVs."
        )
    livro = load_workbook(filename=caminho, read_only=True, data_only=True)
    tabelas, achados = {}, []
    presentes = list(livro.sheetnames)
    esperadas = {c["aba"]: n for n, c in CONTRATO.items()}
    for aba, nome in esperadas.items():
        if aba not in presentes:
            achados.append(_achado("FONTE", "tabela", nome, "aviso",
                                   f"Aba esperada ausente na planilha: {aba}",
                                   "observado", "O negócio confirma se a aba foi descontinuada ou se o arquivo veio incompleto"))
            continue
        planilha = livro[aba]
        linhas = [[_texto(c) for c in linha] for linha in planilha.iter_rows(values_only=True)]
        linhas = [linha for linha in linhas if any(v != "" for v in linha)]
        cabecalho = linhas[0] if linhas else []
        registros = [dict(zip(cabecalho, linha + [""] * (len(cabecalho) - len(linha)))) for linha in linhas[1:]]
        tabelas[nome] = {"colunas": cabecalho, "registros": registros}
    for aba in presentes:
        if aba not in esperadas:
            achados.append(_achado("FONTE", "aba", aba, "aviso",
                                   f"Aba não prevista no contrato da fonte: {aba} — ignorada pelo diagnóstico",
                                   "observado", "O negócio informa se a aba passa a fazer parte da rotina"))
    livro.close()
    hash_arquivo = _sha256(caminho)
    fontes = {nome: {"origem": f"{caminho.name}::{CONTRATO[nome]['aba']}", "sha256": hash_arquivo} for nome in tabelas}
    return tabelas, fontes, achados


def _sha256(caminho: Path) -> str:
    digest = hashlib.sha256()
    with caminho.open("rb") as fh:
        for bloco in iter(lambda: fh.read(65536), b""):
            digest.update(bloco)
    return digest.hexdigest()


# ---------------------------------------------------------------------------
# Achados
# ---------------------------------------------------------------------------
def _achado(classe, entidade, ident, severidade, descricao, evidencia_status, decisao_pendente, evidencia=""):
    return {
        "classe": classe,
        "entidade": entidade,
        "id": ident,
        "severidade": severidade,
        "descricao": descricao,
        "status_evidencia": evidencia_status,
        "evidencia": evidencia,
        "decisao_pendente": decisao_pendente,
    }


def perfil_coluna(valores: list[str]) -> dict:
    preenchidos = [v for v in valores if v.strip() != ""]
    tipos = set()
    for v in preenchidos:
        alvo = v.strip()
        if RE_DATAHORA.match(alvo):
            tipos.add("datahora")
        elif RE_DATA.match(alvo):
            tipos.add("data")
        elif RE_INTEIRO.match(alvo):
            tipos.add("inteiro")
        elif RE_DECIMAL.match(alvo):
            tipos.add("decimal")
        else:
            tipos.add("texto")
    if not tipos:
        tipo = "vazio"
    elif tipos == {"inteiro", "decimal"}:
        tipo = "numérico"
    elif len(tipos) == 1:
        tipo = next(iter(tipos))
    else:
        tipo = "misto(" + "+".join(sorted(tipos)) + ")"
    return {
        "tipo_inferido": tipo,
        "registros": len(valores),
        "preenchidos": len(preenchidos),
        "nulos": len(valores) - len(preenchidos),
        "distintos": len(set(preenchidos)),
        "exemplos": sorted(set(preenchidos))[:3],
    }


def _numero(valor: str):
    alvo = valor.strip().replace(",", ".")
    try:
        return float(alvo)
    except ValueError:
        return None


def analisar_tabela(nome: str, tabela: dict) -> tuple[dict, list]:
    contrato = CONTRATO[nome]
    registros, colunas = tabela["registros"], tabela["colunas"]
    achados = []

    faltando = [c for c in contrato["colunas"] if c not in colunas]
    excedentes = [c for c in colunas if c not in contrato["colunas"]]
    if faltando:
        achados.append(_achado("ESQUEMA", nome, "-", "aviso",
                               "Colunas esperadas ausentes: " + ", ".join(faltando),
                               "observado", "O negócio ou a TI confirma a mudança de layout da extração"))
    if excedentes:
        achados.append(_achado("ESQUEMA", nome, "-", "aviso",
                               "Colunas não previstas no contrato: " + ", ".join(excedentes),
                               "observado", "O negócio informa se as colunas novas entram no cálculo"))
    if len(set(colunas)) != len(colunas):
        repetidas = sorted({c for c in colunas if colunas.count(c) > 1})
        achados.append(_achado("ESQUEMA", nome, "-", "anomalia",
                               "Colunas com nome repetido: " + ", ".join(repetidas),
                               "observado", "Corrigir a extração na origem"))

    perfil = {c: perfil_coluna([r.get(c, "") for r in registros]) for c in colunas}

    # completude
    for coluna in contrato["essenciais"]:
        if coluna not in colunas:
            continue
        vazios = [_ident(nome, r, i) for i, r in enumerate(registros, start=2) if r.get(coluna, "").strip() == ""]
        if vazios:
            achados.append(_achado("COMPLETUDE", nome, ", ".join(vazios), "anomalia",
                                   f"Campo essencial '{coluna}' vazio em {len(vazios)} registro(s)",
                                   "observado",
                                   "O negócio decide o tratamento: excluir, corrigir na origem ou bloquear a entrega",
                                   f"linhas: {', '.join(vazios)}"))
    for coluna in colunas:
        if coluna in contrato["essenciais"]:
            continue
        p = perfil[coluna]
        if p["nulos"] and p["preenchidos"]:
            achados.append(_achado("COMPLETUDE", nome, coluna, "informativo",
                                   f"Coluna '{coluna}' com {p['nulos']} valor(es) vazio(s) de {p['registros']}",
                                   "observado", "O negócio confirma se o vazio é legítimo nesta coluna"))
        elif p["nulos"] and not p["preenchidos"]:
            achados.append(_achado("COMPLETUDE", nome, coluna, "aviso",
                                   f"Coluna '{coluna}' inteiramente vazia",
                                   "observado", "O negócio confirma se a coluna ainda é alimentada na origem"))

    # duplicidade de identificador
    chave = contrato["chave"]
    if all(c in colunas for c in chave):
        indice = {}
        for i, r in enumerate(registros, start=2):
            valor = "|".join(r.get(c, "") for c in chave)
            indice.setdefault(valor, []).append(i)
        for valor, linhas in sorted(indice.items()):
            if len(linhas) > 1:
                achados.append(_achado("DUPLICIDADE", nome, valor, "anomalia",
                                       f"Identificador {'+'.join(chave)} repetido em {len(linhas)} registros "
                                       "— nenhuma versão foi descartada por este diagnóstico",
                                       "observado",
                                       "O negócio define qual versão prevalece e como registrar o descarte",
                                       f"linhas: {', '.join(str(n) for n in linhas)}"))

    # tipos e valores
    for coluna in contrato["numericas"]:
        if coluna not in colunas:
            continue
        invalidos, negativos = [], []
        for i, r in enumerate(registros, start=2):
            bruto = r.get(coluna, "")
            if bruto.strip() == "":
                continue
            numero = _numero(bruto)
            if numero is None:
                invalidos.append(_ident(nome, r, i))
            elif numero < 0:
                negativos.append(_ident(nome, r, i))
        if invalidos:
            achados.append(_achado("CONSISTENCIA", nome, ", ".join(invalidos), "anomalia",
                                   f"Coluna numérica '{coluna}' com valor não numérico",
                                   "observado", "Corrigir na origem ou definir a conversão aprovada"))
        if negativos:
            achados.append(_achado("CONSISTENCIA", nome, ", ".join(negativos), "aviso",
                                   f"Coluna numérica '{coluna}' com valor negativo",
                                   "observado", "O negócio confirma se valor negativo é legítimo, por exemplo um estorno"))
    for coluna in contrato["datas"]:
        if coluna not in colunas:
            continue
        invalidos = [
            _ident(nome, r, i) for i, r in enumerate(registros, start=2)
            if r.get(coluna, "").strip() != ""
            and not (RE_DATA.match(r[coluna].strip()) or RE_DATAHORA.match(r[coluna].strip()))
        ]
        if invalidos:
            achados.append(_achado("CONSISTENCIA", nome, ", ".join(invalidos), "anomalia",
                                   f"Coluna de data '{coluna}' com formato não reconhecido",
                                   "observado", "Corrigir na origem ou definir o formato aprovado"))

    # identificadores fora do padrao (sem normalizar)
    for coluna in [c for c in colunas if c.endswith("_id")]:
        fora = []
        for i, r in enumerate(registros, start=2):
            bruto = r.get(coluna, "")
            if bruto.strip() == "":
                continue
            if bruto != bruto.strip() or not RE_ID_PADRAO.match(bruto.strip()):
                fora.append((_ident(nome, r, i), i, bruto))
        if fora:
            achados.append(_achado("CONSISTENCIA", nome, ", ".join(ident for ident, _, _ in fora), "anomalia",
                                   f"Identificador '{coluna}' fora do padrão observado (espaços ou caixa divergente); "
                                   "valor PRESERVADO como está na fonte",
                                   "observado",
                                   DECISAO_NORMALIZACAO_IDS,
                                   "; ".join(f"linha {linha}: {coluna}={bruto!r}" for _, linha, bruto in fora)))

    # colunas constantes
    for coluna in colunas:
        p = perfil[coluna]
        if p["preenchidos"] == p["registros"] and p["distintos"] == 1 and p["registros"] > 1:
            achados.append(_achado("FONTE", nome, coluna, "aviso",
                                   f"Coluna '{coluna}' com valor único em todos os registros ({p['exemplos'][0]})",
                                   "observado", "O negócio confirma se a coluna é alimentada de fato"))

    # distribuicao de categorias (informativo, sem julgar dominio)
    distribuicoes = {}
    for coluna in contrato["categoricas"]:
        if coluna in colunas:
            contagem = {}
            for r in registros:
                contagem[r.get(coluna, "")] = contagem.get(r.get(coluna, ""), 0) + 1
            distribuicoes[coluna] = dict(sorted(contagem.items()))

    # inventario de status: cada valor observado e uma decisao de tratamento
    # pendente. O diagnostico NAO decide o que entra ou sai do calculo. Sai
    # como informativo (perfil da fonte), nunca como problema.
    if nome != "parametros":
        for coluna in [c for c in contrato["categoricas"] if c == "status" or c.startswith("status")]:
            if coluna not in colunas:
                continue
            ocorrencias = {}
            for i, r in enumerate(registros, start=2):
                ocorrencias.setdefault(r.get(coluna, "").strip(), []).append(_ident(nome, r, i))
            for valor, ids in sorted(ocorrencias.items()):
                amostra = ", ".join(ids[:10]) + (" …" if len(ids) > 10 else "")
                achados.append(_achado(
                    "CONSISTENCIA", nome, valor or "(vazio)", "informativo",
                    f"Status '{valor or '(vazio)'}' presente em {len(ids)} de {len(registros)} registro(s); "
                    "nenhum registro foi excluído por causa do status",
                    "observado",
                    f"O negócio define quais valores de '{coluna}' entram no cálculo e quais são excluídos",
                    f"ids: {amostra}"))

    resumo = {
        "aba_ou_arquivo": CONTRATO[nome]["aba"],
        "registros": len(registros),
        "colunas_lidas": colunas,
        "chave_declarada": chave,
        "perfil_colunas": perfil,
        "distribuicoes": distribuicoes,
    }
    return resumo, achados


def _ident(nome: str, registro: dict, linha: int) -> str:
    chave = CONTRATO[nome]["chave"][0]
    valor = registro.get(chave, "").strip()
    return valor if valor else f"linha {linha}"


# ---------------------------------------------------------------------------
# Relacionamentos e contradicoes entre tabelas
# ---------------------------------------------------------------------------
def analisar_relacionamentos(tabelas: dict) -> tuple[list, list]:
    resultados, achados = [], []
    for rel in RELACIONAMENTOS:
        (t_org, c_org), (t_dst, c_dst) = rel["origem"], rel["destino"]
        if t_org not in tabelas or t_dst not in tabelas:
            continue
        destino = {r.get(c_dst, "") for r in tabelas[t_dst]["registros"]}
        destino_normalizado = {v.strip().upper(): v for v in destino}
        sem_correspondencia = []
        for i, r in enumerate(tabelas[t_org]["registros"], start=2):
            valor = r.get(c_org, "")
            if valor.strip() == "" or valor in destino:
                continue
            candidato = destino_normalizado.get(valor.strip().upper())
            sem_correspondencia.append({
                "id_origem": _ident(t_org, r, i),
                "valor": valor,
                "linha": i,
                "candidato_apos_normalizacao": candidato,
            })
        resultados.append({
            "relacao": f"{t_org}.{c_org} -> {t_dst}.{c_dst}",
            "forca": rel["forca"],
            "registros_origem": len(tabelas[t_org]["registros"]),
            "sem_correspondencia": len(sem_correspondencia),
            "detalhe": sem_correspondencia,
        })
        for item in sem_correspondencia:
            if item["candidato_apos_normalizacao"]:
                achados.append(_achado(
                    "RELACIONAMENTO", t_org, item["id_origem"],
                    "anomalia" if rel["forca"] == "obrigatoria" else "informativo",
                    f"{c_org}={item['valor']!r} não casa com {t_dst}.{c_dst} como está na fonte; "
                    f"casaria com {item['candidato_apos_normalizacao']!r} se normalizado — normalização NÃO aplicada",
                    "hipotese",
                    DECISAO_NORMALIZACAO_IDS,
                    f"linha {item['linha']}"))
            else:
                achados.append(_achado(
                    "RELACIONAMENTO", t_org, item["id_origem"],
                    "anomalia" if rel["forca"] == "obrigatoria" else "informativo",
                    f"{c_org}={item['valor']!r} sem correspondência em {t_dst}.{c_dst} "
                    f"({'quebra de relacionamento obrigatório' if rel['forca'] == 'obrigatoria' else 'ausência informativa'}); "
                    "registro PRESERVADO",
                    "observado",
                    "O negócio decide entre completar o cadastro na origem, excluir o registro ou bloquear a entrega",
                    f"linha {item['linha']}"))
    return resultados, achados


def analisar_contradicoes(tabelas: dict) -> list:
    achados = []

    if "visitas" in tabelas:
        for i, r in enumerate(tabelas["visitas"]["registros"], start=2):
            status = r.get("status", "").strip()
            realizada = r.get("data_realizada", "").strip()
            planejada = r.get("data_planejada", "").strip()
            vid = r.get("visita_id", "").strip() or f"linha {i}"
            if status.lower().startswith("realizad") and realizada == "":
                achados.append(_achado("CONSISTENCIA", "visitas", vid, "anomalia",
                                       f"Status '{status}' sem data_realizada — status e data se contradizem",
                                       "observado",
                                       "O negócio decide se a visita conta como realizada e sob qual evidência",
                                       f"linha {i}"))
            if status != "" and not status.lower().startswith("realizad") and realizada != "":
                achados.append(_achado("CONSISTENCIA", "visitas", vid, "anomalia",
                                       f"Status '{status}' com data_realizada preenchida — status e data se contradizem",
                                       "observado", "O negócio decide qual campo prevalece", f"linha {i}"))
            if realizada and planejada and RE_DATA.match(realizada) and RE_DATA.match(planejada) and realizada < planejada:
                achados.append(_achado("CONSISTENCIA", "visitas", vid, "aviso",
                                       "Data realizada anterior à data planejada",
                                       "observado", "O negócio confirma se a antecipação é válida", f"linha {i}"))
            mes = r.get("mes_ref", "").strip()
            if mes and planejada and RE_DATA.match(planejada) and not planejada.startswith(mes):
                achados.append(_achado("CONSISTENCIA", "visitas", vid, "aviso",
                                       f"mes_ref {mes} não corresponde à data_planejada {planejada}",
                                       "observado", "O negócio confirma qual campo define o mês de competência", f"linha {i}"))

    if "vendas" in tabelas:
        for i, r in enumerate(tabelas["vendas"]["registros"], start=2):
            pid = r.get("pedido_id", "").strip() or f"linha {i}"
            receita, desconto = _numero(r.get("receita_bruta", "")), _numero(r.get("desconto", ""))
            if receita is not None and desconto is not None and desconto > receita:
                achados.append(_achado("CONSISTENCIA", "vendas", pid, "anomalia",
                                       "Desconto maior que a receita bruta",
                                       "observado", "O negócio confirma se o registro é válido", f"linha {i}"))

    # status entre tabelas: cliente com status distinto de 'Ativo' movimentado
    if "clientes" in tabelas:
        status_cliente = {r.get("cliente_id", "").strip(): r.get("status", "").strip()
                          for r in tabelas["clientes"]["registros"]}
        for tabela, coluna in (("vendas", "cliente_id"), ("visitas", "cliente_id")):
            if tabela not in tabelas:
                continue
            for i, r in enumerate(tabelas[tabela]["registros"], start=2):
                cid = r.get(coluna, "").strip()
                status = status_cliente.get(cid)
                if status and status.lower() != "ativo":
                    achados.append(_achado("CONSISTENCIA", tabela, _ident(tabela, r, i), "aviso",
                                           f"Movimento vinculado a cliente {cid} com status '{status}' no cadastro",
                                           "observado",
                                           "O negócio decide se cliente não ativo entra no relatório",
                                           f"linha {i}"))

    if "parametros" in tabelas:
        for i, r in enumerate(tabelas["parametros"]["registros"], start=2):
            status = r.get("status", "").strip()
            if status and status.lower() != "vigente":
                achados.append(_achado("FONTE", "parametros", r.get("parametro", "").strip() or f"linha {i}", "aviso",
                                       f"Parâmetro com status '{status}' — valor ainda não é regra vigente",
                                       "observado",
                                       "O negócio valida o parâmetro antes de usá-lo em número entregue",
                                       f"linha {i}; fonte declarada: {r.get('fonte', '')}"))

    return achados


def analisar_periodo(tabelas: dict, periodo: str | None) -> list:
    achados = []
    observados = []
    for nome, tabela in tabelas.items():
        for coluna in CONTRATO[nome]["datas"] + (["mes_ref"] if "mes_ref" in tabela["colunas"] else []):
            for r in tabela["registros"]:
                valor = r.get(coluna, "").strip()
                if RE_DATA.match(valor) or RE_DATAHORA.match(valor) or RE_COMPETENCIA.match(valor):
                    observados.append((valor[:7], nome, coluna, r))
    if not observados:
        return achados
    meses = sorted({m for m, _, _, _ in observados})
    achados.append(_achado("FONTE", "fonte", "-", "informativo",
                           f"Competências observadas na fonte: {meses[0]} a {meses[-1]} ({len(meses)} mês(es))",
                           "observado", "-"))
    if periodo:
        inicio, fim = (periodo.split(":") + [periodo])[:2]
        fora = sorted({f"{nome}.{coluna}={mes}" for mes, nome, coluna, _ in observados if mes < inicio or mes > fim})
        if fora:
            achados.append(_achado("FONTE", "fonte", "-", "aviso",
                                   f"Datas fora do período informado ({inicio} a {fim}): {len(fora)} ocorrência(s)",
                                   "observado",
                                   "O negócio confirma se o arquivo mensal veio com competência extra",
                                   "; ".join(fora[:10])))
    return achados


# ---------------------------------------------------------------------------
# Relatorios
# ---------------------------------------------------------------------------
ORDEM_SEVERIDADE = {"anomalia": 0, "aviso": 1, "informativo": 2}


def montar_payload(rotulo, entrada, tipo_entrada, fontes, tabelas, periodo, achados_iniciais=()):
    resumos, achados = {}, list(achados_iniciais)
    for nome in sorted(tabelas):
        resumo, achados_tabela = analisar_tabela(nome, tabelas[nome])
        resumo["fonte"] = fontes.get(nome, {})
        resumos[nome] = resumo
        achados.extend(achados_tabela)
    relacionamentos, achados_rel = analisar_relacionamentos(tabelas)
    achados.extend(achados_rel)
    achados.extend(analisar_contradicoes(tabelas))
    achados.extend(analisar_periodo(tabelas, periodo))

    def ordenar(itens):
        return sorted(itens, key=lambda a: (ORDEM_SEVERIDADE[a["severidade"]], a["classe"],
                                            a["entidade"], str(a["id"])))

    # Duas listas separadas: o que exige atencao (anomalia/aviso) e o perfil da
    # fonte (informativo). O inventario nao infla a contagem de problemas.
    atencao = ordenar([a for a in achados if a["severidade"] in ("anomalia", "aviso")])
    perfil_fonte = ordenar([a for a in achados if a["severidade"] == "informativo"])
    for i, achado in enumerate(atencao, start=1):
        achado["codigo"] = f"D-{i:03d}"
    for i, achado in enumerate(perfil_fonte, start=1):
        achado["codigo"] = f"P-{i:03d}"

    por_classe, por_severidade = {}, {}
    for a in atencao:
        por_classe[a["classe"]] = por_classe.get(a["classe"], 0) + 1
        por_severidade[a["severidade"]] = por_severidade.get(a["severidade"], 0) + 1

    return {
        "versao_diagnostico": VERSAO,
        "natureza": "observacional — nenhuma regra de tratamento aplicada, nenhum indicador calculado",
        "rotulo": rotulo,
        "entrada": {"nome": Path(entrada).name, "tipo": tipo_entrada, "periodo_informado": periodo},
        "tabelas": resumos,
        "relacionamentos": relacionamentos,
        "achados": atencao,
        "perfil_fonte": perfil_fonte,
        "resumo": {
            "tabelas_lidas": len(resumos),
            "registros_lidos": {n: r["registros"] for n, r in resumos.items()},
            "achados_total": len(atencao),
            "achados_por_classe": dict(sorted(por_classe.items())),
            "achados_por_severidade": dict(sorted(por_severidade.items(), key=lambda kv: ORDEM_SEVERIDADE[kv[0]])),
            "itens_perfil_fonte": len(perfil_fonte),
            "decisoes_pendentes": sorted({a["decisao_pendente"] for a in atencao if a["decisao_pendente"] != "-"}),
        },
    }


def render_markdown(payload: dict) -> str:
    linhas = [
        f"# Diagnóstico de qualidade da fonte — {payload['rotulo']}",
        "",
        f"> Diagnóstico **observacional** (v{payload['versao_diagnostico']}). Descreve a fonte como ela está: "
        "não normaliza, não deduplica, não exclui, não imputa, não calcula indicador e não altera o arquivo de origem. "
        "Cada achado aponta a decisão que o negócio precisa tomar; nenhuma anomalia aqui é regra aprovada.",
        "",
        "## Fonte",
        "",
        f"- Entrada: `{payload['entrada']['nome']}` ({payload['entrada']['tipo']})",
        f"- Período informado: {payload['entrada']['periodo_informado'] or 'não informado'}",
        "",
        "| Tabela | Origem | SHA-256 (entrada) | Registros |",
        "| --- | --- | --- | --- |",
    ]
    for nome, resumo in payload["tabelas"].items():
        fonte = resumo.get("fonte", {})
        linhas.append(f"| {nome} | `{fonte.get('origem', '-')}` | `{fonte.get('sha256', '-')[:16]}…` | {resumo['registros']} |")

    resumo = payload["resumo"]
    linhas += [
        "",
        "## Resumo",
        "",
        f"- Tabelas lidas: **{resumo['tabelas_lidas']}**",
        f"- Achados que exigem atenção: **{resumo['achados_total']}** — "
        + (", ".join(f"{k}: {v}" for k, v in resumo["achados_por_severidade"].items()) or "nenhum"),
        f"- Por classe: " + (", ".join(f"{k}: {v}" for k, v in resumo["achados_por_classe"].items()) or "—"),
        f"- Itens de perfil/inventário da fonte (sem juízo, não são problemas): **{resumo['itens_perfil_fonte']}**",
        "",
        "## Esquema descoberto",
        "",
    ]
    for nome, res in payload["tabelas"].items():
        linhas += [
            f"### {nome} (`{res['aba_ou_arquivo']}`) — {res['registros']} registros, "
            f"chave declarada: {'+'.join(res['chave_declarada'])}",
            "",
            "| Coluna | Tipo inferido | Preenchidos | Vazios | Distintos | Exemplos |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
        for coluna in res["colunas_lidas"]:
            p = res["perfil_colunas"][coluna]
            exemplos = ", ".join(f"`{e}`" for e in p["exemplos"]) or "—"
            linhas.append(f"| {coluna} | {p['tipo_inferido']} | {p['preenchidos']} | {p['nulos']} | {p['distintos']} | {exemplos} |")
        for coluna, dist in res["distribuicoes"].items():
            linhas.append("")
            linhas.append(f"Distribuição de `{coluna}`: "
                          + ", ".join(f"{k or '(vazio)'} = {v}" for k, v in dist.items()))
        linhas.append("")

    linhas += ["## Relacionamentos", "", "| Relação | Força | Registros | Sem correspondência |", "| --- | --- | --- | --- |"]
    for rel in payload["relacionamentos"]:
        linhas.append(f"| `{rel['relacao']}` | {rel['forca']} | {rel['registros_origem']} | {rel['sem_correspondencia']} |")

    linhas += ["", "## Achados que exigem atenção", "",
               "Anomalias e avisos: algo contradiz a estrutura esperada, impede um cruzamento confiável "
               "ou precisa de confirmação antes de virar número. Nenhum deles é regra aprovada.",
               "",
               "| Código | Sev. | Classe | Entidade | ID | Descrição | Evidência | Decisão pendente |",
               "| --- | --- | --- | --- | --- | --- | --- | --- |"]
    if payload["achados"]:
        for a in payload["achados"]:
            linhas.append(
                f"| {a['codigo']} | {a['severidade']} | {a['classe']} | {a['entidade']} | {a['id']} | "
                f"{a['descricao']} | {a['evidencia'] or '—'} | {a['decisao_pendente']} |"
            )
    else:
        linhas.append("| — | — | — | — | — | nenhum achado de atenção nesta execução | — | — |")

    linhas += ["", "## Perfil e inventário da fonte", "",
               "Retrato da fonte, sem juízo: **não são problemas** e não entram na contagem de achados. "
               "Servem para o negócio ver o que a base contém (inventário de status, competências, "
               "colunas com vazios legítimos).",
               "",
               "| Código | Classe | Entidade | Item | Descrição | Evidência | Decisão pendente |",
               "| --- | --- | --- | --- | --- | --- | --- |"]
    if payload["perfil_fonte"]:
        for a in payload["perfil_fonte"]:
            linhas.append(
                f"| {a['codigo']} | {a['classe']} | {a['entidade']} | {a['id']} | "
                f"{a['descricao']} | {a['evidencia'] or '—'} | {a['decisao_pendente']} |"
            )
    else:
        linhas.append("| — | — | — | — | nenhum item de perfil nesta execução | — | — |")

    if resumo["decisoes_pendentes"]:
        linhas += ["", "## Decisões pendentes do negócio", "",
                   "Derivadas apenas dos achados de atenção:", ""]
        linhas += [f"- {d}" for d in resumo["decisoes_pendentes"]]
    linhas += ["", "---", "",
               "Nenhum registro foi corrigido, excluído ou completado por este diagnóstico; "
               "as regras de tratamento são decididas pelo negócio e implementadas em uma mudança posterior.", ""]
    return "\n".join(linhas)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def executar(entrada: Path, saida: Path, rotulo: str | None, periodo: str | None) -> dict:
    if entrada.is_dir():
        tabelas, fontes, achados_leitura = ler_csvs(entrada)
        tipo = "pasta com CSVs (uma por aba)"
    elif entrada.suffix.lower() in (".xlsx", ".xlsm"):
        tabelas, fontes, achados_leitura = ler_xlsx(entrada)
        tipo = "planilha Excel"
    else:
        raise SystemExit(f"ERRO: entrada não suportada: {entrada} — esperado .xlsx ou pasta com CSVs")
    if not tabelas:
        raise SystemExit("ERRO: nenhuma tabela do contrato foi encontrada na entrada.")

    payload = montar_payload(rotulo or entrada.stem, entrada, tipo, fontes, tabelas, periodo,
                             achados_iniciais=achados_leitura)

    saida.mkdir(parents=True, exist_ok=True)
    base = f"diagnostico_{payload['rotulo']}"
    (saida / f"{base}.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    (saida / f"{base}.md").write_text(render_markdown(payload), encoding="utf-8")
    return payload


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Diagnóstico observacional de qualidade da base operacional: não trata, não calcula.")
    parser.add_argument("--entrada", required=True, help="arquivo .xlsx da base mensal OU pasta com uma CSV por aba")
    parser.add_argument("--saida", default="outputs/diagnostico", help="pasta de saída (padrão: outputs/diagnostico)")
    parser.add_argument("--rotulo", default=None, help="rótulo da execução usado no nome dos arquivos (ex.: 2026-01)")
    parser.add_argument("--periodo", default=None, help="competência esperada, ex.: 2026-01:2026-03")
    args = parser.parse_args(argv)

    payload = executar(Path(args.entrada), Path(args.saida), args.rotulo, args.periodo)
    resumo = payload["resumo"]
    print(f"Diagnóstico v{VERSAO} — {payload['rotulo']}")
    print(f"  tabelas lidas: {resumo['tabelas_lidas']} | registros: {resumo['registros_lidos']}")
    print(f"  achados que exigem atenção: {resumo['achados_total']} ({resumo['achados_por_severidade']})")
    print(f"  itens de perfil/inventário da fonte: {resumo['itens_perfil_fonte']} (não são problemas)")
    print(f"  saida: {Path(args.saida) / ('diagnostico_' + payload['rotulo'])}.md / .json")
    print("  nenhum registro foi tratado, excluído ou corrigido — diagnóstico observacional")
    return 0


if __name__ == "__main__":
    sys.exit(main())
